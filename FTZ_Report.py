import os
import sys
import re
import io
from io import StringIO
from io import BytesIO as StringIO
import time
import datetime
from datetime import date
from datetime import timedelta
import email
import pyzmail
import imaplib
import imapclient
from email.utils import parseaddr
from pathlib import Path
import win32com.client
import codecs
import PyPDF2
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from email.header import decode_header
import bs4
import base64
from bs4 import BeautifulSoup
import ast
import requests
from io import BytesIO
from imapclient import IMAPClient
import tkinter
from tkinter import * 
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from tkinter.ttk import Progressbar
import tkinter as tk
import pandas as pd
import tkinter
from tkinter import * 
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from tkinter.ttk import Progressbar
import tkinter as tk
import itertools
from itertools import chain
from pdfminer3.layout import LAParams, LTTextBox
from pdfminer3.pdfpage import PDFPage
from pdfminer3.pdfinterp import PDFResourceManager
from pdfminer3.pdfinterp import PDFPageInterpreter
from pdfminer3.converter import PDFPageAggregator
from pdfminer3.converter import TextConverter
global t

today =date.today()
t = today.strftime("%d-%b-%Y") # da-Mmm-YYY. Used for todays date 
y = today + timedelta(1)
w = y.strftime('%d-%b-%Y') #used for tomorrows date



root = tk.Tk()
newPath = filedialog.askdirectory(parent = root, initialdir = "/tmp")

msgs = [] # holds email as an object
subjects = [] # holds emails subject line
emails = [] # holds emails body
attach_directory = [] # holds path were excel will be created, will be used globally
attach_directory_date = [] #holds paths where pdfs will be stored for daily use
attach_directory_keyword = []
file_path = []

#counryAbi = ['CN','UK','IT','DE','FR','CR','TW','CA','MX']

def extractEmailInformation():
    server = imapclient.IMAPClient('imap-mail.outlook.com', ssl=True, use_uid=True)
    server.login("coder@uscustombroker.com", "Sor44840")
    server_info = server.select_folder("INBOX", readonly=True)
    UIDs = server.search(['SINCE',t,'BEFORE',w,'FROM','gmontemayor@uscustombroker.com'])
    for i in UIDs:
        rawMessage = server.fetch([i],['BODY[]'])
        msg = pyzmail.PyzMessage.factory(rawMessage[i][b'BODY[]'])
        subjectLine = msg.get_subject()
        checkSubjectLine = re.search(r'FTZ_PDF',str(subjectLine))
        checkSubjectLine1 = re.search(r'FTZ_PDF1',str(subjectLine))
        checkSubjectLine2 = re.search(r'FTZ_PDF2',str(subjectLine))
        checkSubjectLine3 = re.search(r'FTZ_PDF3',str(subjectLine))
        checkSubjectLine4 = re.search(r'FTZ_PDF4',str(subjectLine))
        checkSubjectLine5 = re.search(r'FTZ_PDF5',str(subjectLine))
        checkSubjectLine6 = re.search(r'FTZ_PDF6',str(subjectLine))
        checkSubjectLine7 = re.search(r'FTZ_PDF7',str(subjectLine))
        checkSubjectLine8 = re.search(r'FTZ_PDF8',str(subjectLine))
        checkSubjectLine9 = re.search(r'FTZ_PDF9',str(subjectLine))
        checkSubjectLine10 = re.search(r'FTZ_PDF10',str(subjectLine))
        checkSubjectLine11 = re.search(r'FTZ_PDF11',str(subjectLine))
        checkSubjectLine12 = re.search(r'FTZ_PDF12',str(subjectLine))
        if checkSubjectLine != None or checkSubjectLine1 != None or checkSubjectLine2 != None or checkSubjectLine3 != None or checkSubjectLine4 != None or checkSubjectLine5 != None or checkSubjectLine6 != None or checkSubjectLine7 != None or checkSubjectLine8 != None or checkSubjectLine9 != None or checkSubjectLine10 != None or checkSubjectLine11 != None or checkSubjectLine12 != None:
            if msg.text_part != None and msg.html_part == None:
                body = msg.text_part.get_payload().decode(msg.text_part.charset)
            elif msg.text_part == None and msg.html_part != None:
                html_string = msg.html_part.get_payload().decode(msg.html_part.charset)
                #now that you have the html as a string value you need to parse throught that html
                parsed_html = bs4.BeautifulSoup(html_string,'lxml')
                newText = parsed_html.get_text() #now that object parsed_html is in text format it can be called with get_text(). This is the new non html text 
                body = re.sub(r'(<!---[\s\S]*-->)','',newText) #this will substitiue white spaces. THis is the new html text that is know readible. Since html will create allot of spaces this code is used here
            elif msg.text_part == None and msg.html_part == None:
                print("Body consists of other rather than text and html.")
            else:
                body = msg.text_part.get_payload().decode(msg.text_part.charset)

            subjects.append(subjectLine)
            emails.append(body)
            msgs.append(msg)

    print(range(len(msgs)))
    print(msgs) # prints email as an object
    print(range(len(subjects)))
    print(subjects) # prints emails subject line
    print(range(len(emails)))
    print(emails) # prints body of email
extractEmailInformation()

def folderMaker():
    for i in subjects:
        keyword = i
        if not os.path.exists(newPath + "\\" + "FTZ"):
            os.makedirs(newPath + "\\" + "FTZ") 
        attach_directory.append(newPath + "\\" + "FTZ") # static one and period
        if not os.path.exists(newPath + "\\" + "FTZ" + "\\" + t):
            os.makedirs(newPath + "\\" + "FTZ" + "\\" + t)
        attach_directory_date.append(newPath + "\\" + "FTZ" + "\\" + t) # can create one a day
        if not os.path.exists(newPath + "\\" + "FTZ" + "\\" + t + "\\" + keyword):
            os.makedirs(newPath + "\\" + "FTZ" + "\\" + t + "\\" + keyword)
        attach_directory_keyword.append(newPath + "\\" + "FTZ" + "\\" + t + "\\" + keyword)
folderMaker()

def getAttachments():
    for i in range(len(msgs)):
        for part in msgs[i].walk():
            if part.get_content_maintype() == "multipart":
                continue
            filename = part.get_filename()
            if filename != None and filename.endswith('.pdf'):
                #print(filename)
                filePath = os.path.join(attach_directory_keyword[i],filename) 
                with open(filePath,'wb') as f:
                    f.write(part.get_payload(decode=True))
                    f.close()
getAttachments()

def extractor():
    count = 0
    wb = xlsxwriter.Workbook(newPath+"\\"+"FTZ"+"\\"+t+"\\"+"FTZ.xlsx")
    sheet = wb.add_worksheet() # sheet created to input all info
    for f in range(len(msgs)):
        files = os.listdir(attach_directory_keyword[f])
        print(files)
        for f2 in files:
            print(f2)
            filePath = os.path.join(attach_directory_keyword[f],f2)
            print(filePath)
            resource_manager = PDFResourceManager()
            fake_file_handle = io.StringIO()
            converter = TextConverter(resource_manager, fake_file_handle, laparams=LAParams())
            page_interpreter = PDFPageInterpreter(resource_manager, converter)
            with open(filePath,'rb') as fh:
                for page in PDFPage.get_pages(fh, caching=True, check_extractable=True):
                    page_interpreter.process_page(page)
                text = fake_file_handle.getvalue()
            converter.close()
            fake_file_handle.close()
            #############################
            print(text)
            clean_space = re.sub('\s+','',str(text))
            single_space = re.sub('\s+',' ',str(text))
            print(single_space)
            print("___________________________________________________")
            #############################
            zone_admi = [] #col b
            quantity = [] #col d
            unit = [] #col f
            country = [] #col j
            description = [] #col g
            value = [] #col i
            hmf_duty = [] #col k
            cutDate = [] # col Q
            #############################
            zoneAdmi = re.search(r'(?<=6. ZONE ADMISSION NO. )(.*?)\s+',single_space)
            oneSwip = re.search(r'(?<=CHGS.)\s*[0-9]*\s+[A-Z]*\s+[A-Z]*\s+(.*)',text) #gets col d,f,j and g
            oneSwip2 = re.search(r'(?<=ORIGIN CODE)\s*[0-9]*\s+[A-Z]*\s+[A-Z]*\s+(.*)',text)
            value1 = re.search(r'LBS\s*(.*)\s*21.\s+HARBOR',text) # single value
            value2 = re.search(r'LBS\s*(.*)\s*(.*)\s*21.\s+HARBOR',text) # double value
            value3 = re.search(r'LBS\s*(.*)\s*23.\s+I\s+hereby',text) # single value
            value4 = re.search(r'LBS\s*(.*)\s*(.*)\s*23.\s+I\s+hereby',text) # double value
            duty = re.search(r'\(19\s+CFR\s+24\.24\)\s*(.*)',text)
            date = re.search(r'(?<=30. DATE )[0-9/]*\s+',single_space)
            ##############################
            if zoneAdmi != None: #detected for all
                print(zoneAdmi.group(0))
                zone_admi.append(zoneAdmi.group(0)) #right
            ##############################
            if oneSwip != None: 
                print(oneSwip.group(0))
                split = oneSwip.group(0).split(" ")
                print(split)
                quantity.append(split[0])
                unit.append(split[1])
                country.append(split[2])
                dis = " ".join(split[3:])
                description.append(dis)
            elif oneSwip2 != None:
                print(oneSwip2.group(0))
                split2 = oneSwip2.group(0).split(" ")
                print(split2)
                quantity.append(split2[0])
                unit.append(split2[1])
                country.append(split2[2])
                dis2 = " ".join(split2[3:])
                description.append(dis2)
            ################################
            if value1 != None:
                print(value1.group(0))
                print(value1.group(1))
                value.append(value1.group(1))
            elif value2 != None:
                print(value2.group(0))
                print(value2.group(1))
                print(value2.group(2))
                vall = re.sub(',','',str(value2.group(1)))
                vall2 = re.sub(',','',str(value2.group(2)))
                val = float(vall)
                val2 = float(vall2)
                summ = val + val2
                print(summ)
                value.append(summ)
            elif value3 != None:
                print(value3.group(0))
                print(value3.group(1))
                value.append(value3.group(1))
            elif value4 != None:
                print(value4.group(0))
                print(value4.group(1))
                print(value4.group(2))
                vall3 = re.sub(',','',str(value4.group(1)))
                vall4 = re.sub(',','',str(value4.group(2)))
                val3 = float(vall3)
                val4 = float(vall4)
                summ2 = val3 + val4
                print(summ2)
                value.append(summ2)
            ##################################
            if duty != None:
                print(duty.group(1))
                hmf_duty.append(duty.group(1))
            ##################################
            if date != None:
                print(date.group(0))
                cutDate.append(date.group(0))
            ####################################
            for row_num, data in enumerate(zone_admi):
                row_num = row_num + count
                sheet.write(row_num, 1, data)

            for row_num, data in enumerate(zone_admi):
                row_num = row_num + count
                sheet.write(row_num, 3, quantity[0])

            for row_num, data in enumerate(zone_admi):
                row_num = row_num + count
                sheet.write(row_num, 5, unit[0])

            for row_num, data in enumerate(zone_admi):
                row_num = row_num + count
                sheet.write(row_num, 6, description[0])

            for row_num, data in enumerate(zone_admi):
                row_num = row_num + count
                sheet.write(row_num, 8, value[0])

            for row_num, data in enumerate(zone_admi):
                row_num = row_num + count
                sheet.write(row_num, 9, country[0])

            for row_num, data in enumerate(zone_admi):
                row_num = row_num + count
                sheet.write(row_num, 10, hmf_duty[0])

            for row_num, data in enumerate(zone_admi):
                row_num = row_num + count
                sheet.write(row_num, 16, cutDate[0])
            count += len(zone_admi) #this adds one at a time

    print(zone_admi)
    print(quantity)
    print(unit)
    print(country)
    print(description)
    print(value)
    print(hmf_duty)
    print(cutDate)
    wb.close()
extractor()

def excel_collector():
    fileName = newPath + "\\" + "FTZ" + "\\" + "FTZ.xlsx"
    wb = openpyxl.load_workbook(fileName)
    ws = wb.active #output file
    ############
    fileName1 = newPath + "\\" + "FTZ" + "\\" + t + "\\" + "FTZ.xlsx"
    wb1 = openpyxl.load_workbook(fileName1)
    ws1 = wb1.active #input file
    ############
    mr = ws.max_row
    mc = ws.max_column
    print(mr)                #output file
    print(mc)
    ############
    mr1 = ws1.max_row
    mc1 = ws1.max_column
    print(mr1)               #input file
    print(mc1)
    ############
    for i in range(1, mr1 + 1):
        for j in range(1, mc1 +1):
            c = ws1.cell(row=i,column=j)
            ws.cell(row=mr+i,column=j).value = c.value
    wb.save(str(fileName))
excel_collector()

def quarter_cut():
    fileName = newPath + "\\" + "FTZ" + "\\" + "FTZ.xlsx"
    wb = openpyxl.load_workbook(fileName)
    ws = wb.active
    ###############
    value_add1 = []
    hmf_sub1 = []
    value_add2 = []
    hmf_sub2 = []
    value_add3 = []
    hmf_sub3 = []
    value_add4 = []
    hmf_sub4 = []
    ###############
    for row in ws.iter_rows(min_row=17,min_col=9,max_col=17):
        if row[-1].value >= "01/00/2020" and row[-1].value < "04/00/2020": #1st quarter
            print("QTR1")
            print(row[-1].value) #date
            print(row[0].value) #value
            print(row[2].value) #HMF DUE
            a = re.sub(',','',row[0].value) #value
            b = re.sub(',','',row[2].value) #HMF DUE
            aa = float(a)
            bb = float(b)
            value_add1.append(aa)
            if bb == 0.0:
                hmf_sub1.append(aa)
        elif row[-1].value >= "04/00/2020" and row[-1].value < "07/00/2020": #2nd quarter
            print("QTR2")
            print(row[-1].value)
            print(row[0].value)
            print(row[2].value)
            a = re.sub(',','',row[0].value) #value
            b = re.sub(',','',row[2].value) #HMF DUE
            aa = float(a)
            bb = float(b)
            value_add2.append(aa)
            if bb == 0.0:
                hmf_sub2.append(aa)
        elif row[-1].value >= "07/00/2020" and row[-1].value < "10/00/2020": #3rd quarter
            print("QTR3")
            print(row[-1].value)
            print(row[0].value)
            print(row[2].value)
            a = re.sub(',','',row[0].value) #value
            b = re.sub(',','',row[2].value) #HMF DUE
            aa = float(a)
            bb = float(b)
            value_add3.append(aa)
            if bb == 0.0:
                hmf_sub3.append(aa)
        elif row[-1].value >= "10/00/2020" and row[-1].value < "01/00/2021": #4th quarter
            print("QTR4")
            print(row[-1].value)
            print(row[0].value)
            print(row[2].value)
            a = re.sub(',','',row[0].value) #value
            b = re.sub(',','',row[2].value) #HMF DUE
            aa = float(a)
            bb = float(b)
            value_add4.append(aa)
            if bb == 0.0:
                hmf_sub4.append(aa)
            else:
                None
    ############### SUM VALUES FROM LIST
    if ws["T4"].value != None:
        t4 = sum(value_add1)
        ws["T4"] = ws["T4"].value + t4

    if ws["T5"].value != None:
        t5 = sum(hmf_sub1)
        ws["T5"] = ws["T5"].value + t5

    if ws["W4"].value != None:
        w4 = sum(value_add2)
        ws["W4"] = ws["W4"].value + w4

    if ws["W5"].value != None:
        w5 = sum(hmf_sub2)
        ws["W5"] = ws["W5"].value + w5

    if ws["T12"].value != None:
        t12 = sum(value_add3)
        ws["T12"] = ws["T12"].value + t12

    if ws["T13"].value != None:
        t13 = sum(hmf_sub3)
        ws["T13"] = ws["T13"].value + t13

    if ws["W12"].value != None:
        w12 = sum(value_add4)
        ws["W12"] = ws["W12"].value + w12

    if ws["W13"].value != None:
        w13 = sum(hmf_sub4)
        ws["W13"] = ws["W13"].value + w13
    wb.save(fileName)
quarter_cut()

